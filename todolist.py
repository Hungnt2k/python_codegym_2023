from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
# Mở ứng dụng chính
def open_main_app():

    # Tạo cửa sổ giao diện
    main_app = Tk()
    main_app.title("Task ToDoList")
    main_app.geometry("800x800")
    main_app.configure(bg="#FFFEC4")
    # Kết nối đến tệp Excel
    workbook = load_workbook("tasks.xlsx")
    sheet = workbook.active

    # Khởi tạo Treeview
    tree = ttk.Treeview(main_app)
    tree["columns"] = ("task_title", "description", "owner", "start_date", "due_date", "dur", "complete", "result")
    tree.column("#0", width=0, stretch=NO)
    tree.column("task_title", width=120)
    tree.column("description", width=120)
    tree.column("owner", width=80)
    tree.column("start_date", width=80)
    tree.column("due_date", width=80)
    tree.column("dur", width=40)
    tree.column("complete", width=60)
    tree.column("result", width=120)

    tree.heading("task_title", text="Task Title")
    tree.heading("description", text="Description")
    tree.heading("owner", text="Owner")
    tree.heading("start_date", text="Start Date")
    tree.heading("due_date", text="Due Date")
    tree.heading("dur", text="Dur")
    tree.heading("complete", text="Complete")
    tree.heading("result", text="Result")
    
    # Đăng xuất
    def logout():
        workbook.save("tasks.xlsx")
        messagebox.showinfo("Logout", "Logout successful!")
        main_app.destroy()
        init_app()

    # Lấy dữ liệu từ tệp Excel và hiển thị trên Treeview
    def load_data():
        for row in sheet.iter_rows(values_only=True):
            tree.insert("", "end", values=row)

    # Thêm task mới
    def add_task():
        title = entry_title.get()
        description = entry_description.get()
        owner = entry_owner.get()
        start_date = entry_start_date.get()
        due_date = entry_due_date.get()
        dur = entry_dur.get()
        complete = entry_complete.get()
        result = entry_result.get()

        if title == "" or description == "" or owner == "":
            messagebox.showwarning("Invalid Input", "Please enter values for Task Title, Description, and Owner.")
            return

        dur = validate_dur(dur)
        if dur is None:
            return
        try:
            start_date = datetime.strptime(start_date, "%d-%m-%Y")
            start_date_str = start_date.strftime("%d-%m-%Y")
            due_date = datetime.strptime(due_date, "%d-%m-%Y")
            due_date_str = due_date.strftime("%d-%m-%Y")
        except ValueError: 
            messagebox.showerror("Invalid Input", "Please enter the date in the format dd-mm-yyyy.")
            return 

        sheet.append([title, description, owner, start_date, due_date, dur, complete, result])
        workbook.save("tasks.xlsx")

        tree.insert("", "end", values=(title, description, owner, start_date_str, due_date_str, dur, complete, result))
        messagebox.showinfo("Success", "Task added successfully!")
        clear_entries()

    # Xóa task được chọn
    def delete_task():
        selected_item = tree.selection()
        if selected_item: 
            task_title = tree.item(selected_item)["values"][0]
            confirmation = messagebox.askyesno("Confirm", f"Are you sure you want to delete the task '{task_title}'?")
            if confirmation:
                for item in selected_item:
                    index = tree.index(item)
                    tree.delete(item)
                    sheet.delete_rows(index + 2)
                workbook.save("tasks.xlsx")
                messagebox.showinfo("Success", "Task deleted successfully!")
                
        else:
            messagebox.showwarning("No Selection", "Please select a task to delete.")

    # Cập nhật task được chọn
    def update_task():
        selected_item = tree.selection()
        if selected_item:
            task_title = tree.item(selected_item)["values"][0]
            confirmation = messagebox.askyesno("Confirm", f"Are you sure you want to update the task '{task_title}'?")
            if confirmation:
                for item in selected_item:
                    index = tree.index(item)
                    task_values = tree.item(item)["values"]
                    try:
                        updated_values = [
                            entry_title.get() if entry_title.get() else task_values[0],
                            entry_description.get() if entry_description.get() else task_values[1],
                            entry_owner.get() if entry_owner.get() else task_values[2],
                            entry_start_date.get() if entry_start_date.get() else task_values[3],
                            entry_due_date.get() if entry_due_date.get() else task_values[4],
                            entry_dur.get() if entry_dur.get() else task_values[5],
                            entry_complete.get() if entry_complete.get() else task_values[6],
                            entry_result.get() if entry_result.get() else task_values[7]
                        ]
                    except IndexError:
                        updated_values = [
                            entry_title.get() if entry_title.get() else task_values[0],
                            entry_description.get() if entry_description.get() else task_values[1],
                            entry_owner.get() if entry_owner.get() else task_values[2],
                            entry_start_date.get() if entry_start_date.get() else "",
                            entry_due_date.get() if entry_due_date.get() else "",
                            entry_dur.get() if entry_dur.get() else "",
                            entry_complete.get() if entry_complete.get() else "",
                            entry_result.get() if entry_result.get() else ""
                        ]
                    
                    dur = validate_dur(updated_values[5])
                    if dur is None:
                        return
                    updated_values[5] = dur
                    try:    
                        start_date = datetime.strptime(updated_values[3], "%d-%m-%Y")
                        start_date_str = start_date.strftime("%d-%m-%Y")
                        due_date = datetime.strptime(updated_values[4], "%d-%m-%Y")
                        due_date_str = due_date.strftime("%d-%m-%Y")
                    except ValueError: 
                        messagebox.showerror("Invalid Input", "Please enter the date in the format dd-mm-yyyy.")
                        return
                    updated_values[3] = start_date_str
                    updated_values[4] = due_date_str
                    for i in range(len(task_values)):
                        if task_values[i] != updated_values[i]:
                            tree.item(item, values=updated_values)
                    for i, value in enumerate(updated_values):
                        sheet.cell(row=index + 2, column=i + 1).value = value
                workbook.save("tasks.xlsx")
                messagebox.showinfo("Success", "Task updated successfully!")
                clear_entries()
        else:
            messagebox.showwarning("No Selection", "Please select a task to update.")


    # Tìm kiếm task theo tiêu đề
    def search_task():
        search_value = entry_search.get().lower()
        for item in tree.get_children():
            task_title = tree.item(item)["values"][0].lower()
            if search_value in task_title:
                tree.selection_set(item)
                tree.focus(item)
                tree.see(item)
                return
        messagebox.showwarning("No Match", "No tasks found matching the search criteria.")
    def reload_data():
    # Xóa toàn bộ các items hiện tại trên Treeview
        tree.delete(*tree.get_children())

    # Mở lại tệp Excel
        workbook = load_workbook("tasks.xlsx")
        sheet = workbook.active

    # Load lại dữ liệu từ tệp Excel và hiển thị trên Treeview
        for row in sheet.iter_rows(values_only=True):
            tree.insert("", "end", values=row)
        messagebox.showinfo("Reload", "Data reloaded successfully!")
    # Xóa nội dung của các trường nhập
    def clear_entries():
        entry_title.delete(0, END)
        entry_description.delete(0, END)
        entry_owner.delete(0, END)
        entry_start_date.delete(0, END)
        entry_due_date.delete(0, END)
        entry_dur.delete(0, END)
        entry_complete.delete(0, END)
        entry_result.delete(0, END)

    def validate_dur(dur):
        if dur.endswith("%"):
            return dur
        else:
            messagebox.showerror("Invalid Input", "Please enter a value with % for Dur.")
        return None
    # Hiển thị tên người đăng nhập và nút Logout
    def show_user_info():
        username = f"Logged in as: {current_user}"
        label_user_info.config(text=username)

        button_logout.pack(side=RIGHT, padx=10)

    # Giao diệnngười dùng
    frame_user = Frame(main_app, bg="#FFFEC4")
    frame_user.pack(pady=10)

    label_user_info = Label(frame_user, text="", bg="#FFFEC4")
    label_user_info.pack(side=LEFT, padx=10)

    button_logout = Button(frame_user, text="Logout", command=logout, bg="#FF5722", fg="white")
    button_logout.pack(side=RIGHT, padx=10)

    # Giao diện chính
    frame_title = Frame(main_app)
    frame_title.pack(pady=10)

    label_app_title = Label(frame_title, text="Task ToDoList", font=("Time News Roman", 18, "bold"), bg="#FFFEC4")
    label_app_title.pack()

    frame_entries = Frame(main_app, bg="#FFFEC4")
    frame_entries.pack(pady=10)

    entry_title = Entry(frame_entries, width=20)
    entry_title.grid(row=0, column=1, padx=5, pady=5)
    label_title = Label(frame_entries, text="Task Title:", font=("Arial", 10), bg="#FFFEC4")
    label_title.grid(row=0, column=0, padx=5, pady=5)

    entry_description = Entry(frame_entries, width=20)
    entry_description.grid(row=1, column=1, padx=5, pady=5)
    label_description = Label(frame_entries, text="Description:", font=("Arial", 10) ,bg="#FFFEC4")
    label_description.grid(row=1, column=0, padx=5, pady=5)

    entry_owner = Entry(frame_entries, width=20)
    entry_owner.grid(row=2, column=1, padx=5, pady=5)
    label_owner = Label(frame_entries, text="Owner:", font=("Arial", 10),bg="#FFFEC4")
    label_owner.grid(row=2, column=0, padx=5, pady=5)

    entry_start_date = Entry(frame_entries, width=20)
    entry_start_date.grid(row=3, column=1, padx=5, pady=5)
    label_start_date = Label(frame_entries, text="Start Date:", font=("Arial", 10),bg="#FFFEC4")
    label_start_date.grid(row=3, column=0, padx=5, pady=5)

    entry_due_date = Entry(frame_entries, width=20)
    entry_due_date.grid(row=4, column=1, padx=5, pady=5)
    label_due_date = Label(frame_entries, text="Due Date:", font=("Arial", 10),bg="#FFFEC4")
    label_due_date.grid(row=4, column=0, padx=5, pady=5)

    entry_dur = Entry(frame_entries, width=20)
    entry_dur.grid(row=5, column=1, padx=5, pady=5)
    label_dur = Label(frame_entries, text="Dur:", font=("Arial", 10),bg="#FFFEC4")
    label_dur.grid(row=5, column=0, padx=5, pady=5)

    entry_complete = Entry(frame_entries, width=20)
    entry_complete.grid(row=6, column=1, padx=5, pady=5)
    label_complete = Label(frame_entries, text="Complete:", font=("Arial", 10),bg="#FFFEC4")
    label_complete.grid(row=6, column=0, padx=5, pady=5)

    entry_result = Entry(frame_entries, width=20)
    entry_result.grid(row=7, column=1, padx=5, pady=5)
    label_result = Label(frame_entries, text="Result:", font=("Arial", 10),bg="#FFFEC4")
    label_result.grid(row=7, column=0, padx=5, pady=5)

    if current_user == "admin":
        button_add = Button(frame_entries, text="Add Task", command=add_task, bg="#4CAF50", fg="white")
        button_add.grid(row=8, column=0, padx=5, pady=5)

        button_delete = Button(frame_entries, text="Delete Task", command=delete_task, bg="#F44336", fg="white")
        button_delete.grid(row=8, column=2, padx=5, pady=5)

    button_update = Button(frame_entries, text="Update Task", command=update_task, bg="#2196F3", fg="white")
    button_update.grid(row=8, column=1, padx=5, pady=5)

    frame_buttons = Frame(main_app, bg="#FFFEC4")
    frame_buttons.pack(pady=10)
    if current_user == "admin":
        button_reload = Button(frame_buttons, text="Reload", command=reload_data, bg="#FFC107", fg="white")
        button_reload.pack(side=LEFT, padx=10)

# Các nút khác (Add Task, Delete Task, Update Task) nằm trong frame_buttons cùng với nút Reload

    frame_search = Frame(main_app, bg="#FFFEC4")
    frame_search.pack(pady=10)

    entry_search = Entry(frame_search, width=20)
    entry_search.grid(row=0, column=1, padx=5, pady=5)
    label_search = Label(frame_search, text="Search:", font=("Arial", 10),bg="#FFFEC4")
    label_search.grid(row=0, column=0, padx=5, pady=5)

    button_search = Button(frame_search, text="Search", command=search_task, bg="#FF9800", fg="white")
    button_search.grid(row=0, column=2, padx=5, pady=5)

    # Hiển thị tên người đăng nhập
    show_user_info()

    # Load dữ liệu từ tệp Excel
    load_data()

    # Hiển thị Treeview
    tree.pack(pady=10)

    # Chạy ứng dụng chính
    main_app.mainloop()


def init_app():
    # Kiểm tra đăng nhập
    def login():
        global logged_in, current_user
        username = entry_username.get()
        password = entry_password.get()

        if username == "admin" and password == "admin":
            messagebox.showinfo("Success", "Login successful!")
            logged_in = True
            current_user = "admin"
            root.destroy()
            # Thực thi ứng dụng chính
            open_main_app()
        elif username == "user" and password == "user":
            messagebox.showinfo("Success", "Login successful!")
            logged_in = True
            current_user = "user"
            root.destroy()
            # Thực thi ứng dụng chính với quyền hạn của người dùng
            open_main_app()
        elif username == "hungnt" and password == "hungnt":
            messagebox.showinfo("Success", "Login successful!")
            logged_in = True
            current_user = "hungnt"
            root.destroy()
            # Thực thi ứng dụng chính với quyền hạn của "hungnt"
            open_main_app()
        else:
            messagebox.showwarning("Invalid Credentials", "Invalid username or password!")

    def handle_enter(event):
        if event.keycode == 13:
            login()        
    # Tạo cửa sổ đăng nhập
    root = Tk()
    root.title("Login Task")
    root.geometry("300x180")
    root.configure(bg="#FFD6A5")
    # Biến lưu trạng thái đăng nhập
    logged_in = False
    current_user = ""
    # Giao diện người dùng
    frame = Frame(root, bg="#FFD6A5")
    frame.pack(pady=20)

    label_app_title = Label(frame, text="Login Task", font=("Arial", 16, "bold"), fg="blue", bg="#FFD6A5")
    label_app_title.grid(row=0, column=0, columnspan=2)

    entry_username = Entry(frame, width=20)
    entry_username.grid(row=1, column=1, padx=5, pady=5)
    label_username = Label(frame, text="Username:", font=("Arial", 10, "bold"), bg="#FFD6A5")
    label_username.grid(row=1, column=0, padx=5, pady=5)
                        
    entry_password = Entry(frame, width=20, show="*")
    entry_password.grid(row=2, column=1, padx=5, pady=5)
    label_password = Label(frame, text="Password:", font=("Arial", 10, "bold"), bg="#FFD6A5")
    label_password.grid(row=2,column=0, padx=5, pady=5)

    button_login = Button(frame, text="Login", command=login, bg="#2196F3", fg="white",  font=("Arial", 11))
    button_login.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

    root.bind('<Key>', handle_enter)
    # Chạy cửa sổ đăng nhập
    root.mainloop()
        

if __name__ == "__main__":
    init_app()
